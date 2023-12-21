import os
import re
import time
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import Workbook, load_workbook

ua = UserAgent()
# 随机生成一个User-Agent
random_user_agent = ua.random
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/54.0.2840.99 Safari/537.36"
        #random_user_agent
}

def get_onpage():
    for i in range(11):
        print(f'正在爬取第{i+1}页,请稍等...')
        url = f'https://movie.douban.com/top250?start={i*25}&filter='

        while True:
            try:
                r = requests.get(url,headers=headers, timeout=10)
                r.raise_for_status()
                r.encoding='utf-8'
                soup=BeautifulSoup(r.text,'html.parser')
                lists=soup.find_all('div', attrs={'class': 'hd'})

                for list in lists:
                    f_url=list.a['href']
                    time.sleep(0.5)
                    get_allinfo(f_url)
                break  # 如果成功爬取并处理了页面，就跳出循环
            except:
                print("error!")


def get_allinfo(url):
    filminfo = []
    r = requests.get(url, headers=headers, timeout=10)
    r.raise_for_status()
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'html.parser')
    # 片名
    name = soup.find(attrs={'property': 'v:itemreviewed'}).text.split(' ')[0]
    # 上映年份
    year = soup.find(attrs={'class': 'year'}).text.replace('(', '').replace(')', '')
    # 评分
    score = soup.find(attrs={'property': 'v:average'}).text
    # 评价人数
    votes = soup.find(attrs={'property': 'v:votes'}).text
    infos = soup.find(attrs={'id': 'info'}).text.split('\n')[1:11]
    # 导演
    director = infos[0].split(': ')[1]
    # 编剧
    scriptwriter = infos[1].split(': ')[1]
    # 主演
    actor = infos[2].split(': ')[1]
    # 类型
    filmtype = infos[3].split(': ')[1]
    # 国家/地区
    area = infos[4].split(': ')[1]
    if '.' in area:
        area = infos[5].split(': ')[1].split(' / ')[0]
        # 语言
        language = infos[6].split(': ')[1].split(' / ')[0]
    else:
        area = infos[4].split(': ')[1].split(' / ')[0]
        # 语言
        language = infos[5].split(': ')[1].split(' / ')[0]

    if '大陆' in area or '香港' in area or '台湾' in area:
        area = '中国'
    if '戛纳' in area:
        area = '法国'
    # 时长
    times0 = soup.find(attrs={'property': 'v:runtime'}).text
    times = re.findall('\d+', times0)[0]
    filminfo.append(name)
    filminfo.append(year)
    filminfo.append(score)
    filminfo.append(votes)
    filminfo.append(director)
    filminfo.append(scriptwriter)
    filminfo.append(actor)
    filminfo.append(filmtype)
    filminfo.append(area)
    filminfo.append(language)
    filminfo.append(times)
    filepath = './information/TOP500.xlsx'
    get_excel(filepath, filminfo)


def get_excel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = ['片名', '上映年份', '评分', '评价人数', '导演', '编剧', '主演', '类型', '国家/地区', '语言',
                          '时长(分钟)']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)
            time.sleep(3)

        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        ws.append(allinfo)
        wb.save(filepath)
        return True
    except:
        return False

if __name__ == '__main__':
    get_onpage()