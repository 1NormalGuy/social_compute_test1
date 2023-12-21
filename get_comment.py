import requests
import time
import dic
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from random import randint
import csv

tunnel = "y525.kdltps.com:15818"

# #request
# def get_html(url,headers,flag):
#     # 使用requests.get获取网页内容
#     response = requests.get(url, headers=headers,proxies=proxies)

#     # 检查请求是否成功（状态码为200）
#     if response.status_code == 200:
#         return response.text
#     else:
#         return None

def get_html(url,headers):
    while True:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.text
        else:
            print(f"请求失败，状态码：{response.status_code}")
            print("等待5秒后重试...")
            time.sleep(5)

def get_comments(url,headers,arr):
    # 使用requests.get获取网页内容
    html=get_html(url,headers)
    if html:
        #收集地区信息
        get_locations(html, arr)
        soup = BeautifulSoup(html, 'lxml')
        comments = []
        comment_divs = soup.find_all('p', {'class': 'comment-content'})
        for comment_div in comment_divs:
            comment = comment_div.find('span', {'class': 'short'})
            if comment:
                comments.append(comment.text)

        return comments
    return None


def get_locations(html,arr):
    if html:
        soup = BeautifulSoup(html, 'lxml')
        locations = []
        locations_divs = soup.find_all('span', {'class': 'comment-location'})
        for locations_div in locations_divs:
            if locations_div:
                locations.append(locations_div.text)

        for ip in locations:
            if ip:
                if ip not in dic.comment_location:
                    ip = '海外'
                arr[dic.comment_location[ip]] = arr[dic.comment_location[ip]] + 1


def save_comments(comments, filename):
    # 使用'a'模式打开文件，这将会创建一个新的文件或续写已存在的文件
    with open(filename, 'a', encoding='utf-8') as f:
        writer = csv.writer(f)
        for comment in comments:
            # 写入一行数据
            writer.writerow([comment])


def get_excel(filepath, allinfo):
    try:
        if not os.path.exists(filepath):
            tableTitle = [" ","河北","山西","辽宁","吉林","黑龙江","江苏","浙江","安徽","福建","江西","山东","河南","湖北","湖南","广东","海南","四川","贵州","云南","陕西","甘肃","青海","台湾"
,"内蒙古","广西","西藏","宁夏","新疆","北京","天津","上海","重庆","香港","澳门",'海外']
            wb = Workbook()
            ws = wb.active
            ws.title = 'sheet1'
            ws.append(tableTitle)
            wb.save(filepath)

        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = 'sheet1'
        ws.append(allinfo)
        wb.save(filepath)
        return True
    except:
        return False
"""
#获取评论用户个人主页url
def get_user_url(html):
    if html:
        soup = BeautifulSoup(html, 'lxml')
        url_list = []

        # 在comment-info中爬取链接 防止爬到无关链接
        links_divs = soup.find_all('span',{'class': 'comment-info'})
        for links_div in links_divs:
            if links_div:
                links=links_div.find_all('a')
                for link in links:
                    url=link.get('href')
                    url_list.append(url)

        #过滤url_list中的无关元素
        url_list = list(filter(lambda url_str: 'http' in url_str, url_list))
        #过滤重复项
        new_url_list = list(dict.fromkeys(url_list))
    return new_url_list


def get_locations(url,headers,arr):
    html = get_html(url, headers,1)
    if html:
        url_list=get_user_url(html)
        for user_url in url_list:
            user_html=get_html(user_url,headers,1)
            if user_html:
                soup = BeautifulSoup(user_html,'lxml')
                user_ip = soup.find('span', {'class': 'ip-location'})
                if user_ip:
                    #分割字符串删掉IP属地： 只保留地区
                    ip = user_ip.text.split('：')[1]
                    if ip not in dic.comment_location:
                        ip = '海外'
                    arr[dic.comment_location[ip]]=arr[dic.comment_location[ip]]+1
"""
